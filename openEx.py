import tkinter as tk
from openpyxl import load_workbook
from tkinter import ttk

def show_excel(file_path):
    # Load the Excel workbook
    workbook = load_workbook(file_path)
    
    # Get the active sheet
    sheet = workbook.active
    
    # Create a Tkinter window
    root = tk.Tk()
    root.title("Excel Viewer")

    # Create a Treeview widget to display the Excel data
    tree = ttk.Treeview(root)

    # Add columns to the Treeview
    for col in sheet.iter_cols(1, sheet.max_column):
        tree.heading(col[0].column, text=col[0].value)
        tree.column(col[0].column, width=100)

    # Add rows to the Treeview
    for row in sheet.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        tree.insert('', 'end', values=values)

    # Add a vertical scrollbar
    scrollbar = ttk.Scrollbar(root, orient='vertical', command=tree.yview)
    tree.configure(yscroll=scrollbar.set)

    # Pack the Treeview and Scrollbar
    tree.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side='right', fill='y')

    # Run the Tkinter main loop
    root.mainloop()

# Example usage
def openExcel(classNum):
    excelFile = "./excel/"+classNum+".xlsx"
    show_excel(excelFile)

